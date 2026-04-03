import openpyxl

# Excel file name
file = "employee_data.xlsx"

# Load Excel file
try:
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
except:
    print("Excel file not found. Check the file name.")
    exit()

# ---------------- ADD EMPLOYEE ----------------
def add_employee():
    name = input("Enter Name: ")
    email = input("Enter Email: ")
    dept = input("Enter Department: ")
    manager = input("Enter Manager Name: ")
    m_email = input("Enter Manager Email: ")

    sheet.append([name, email, dept, manager, m_email])
    wb.save(file)
    print("Employee added successfully")

# ---------------- VIEW EMPLOYEE ----------------
def view_employee():
    print("\nEmployee Records:")
    for row in sheet.iter_rows(values_only=True):
        print(row)

# ---------------- SEARCH EMPLOYEE ----------------
def search_employee():
    name = input("Enter Name to search: ")
    found = False

    for row in sheet.iter_rows(values_only=True):
        if row[0] == name:
            print("Employee Found:", row)
            found = True

    if not found:
        print("Employee not found")

# ---------------- UPDATE EMPLOYEE ----------------
def update_employee():
    name = input("Enter Name to update: ")

    for row in sheet.iter_rows():
        if row[0].value == name:
            print("Enter new details:")
            row[1].value = input("New Email: ")
            row[2].value = input("New Department: ")
            row[3].value = input("New Manager: ")
            row[4].value = input("New Manager Email: ")

            wb.save(file)
            print("Employee updated successfully")
            return

    print("Employee not found")

# ---------------- DELETE EMPLOYEE ----------------
def delete_employee():
    name = input("Enter Name to delete: ")

    for row in sheet.iter_rows():
        if row[0].value == name:
            sheet.delete_rows(row[0].row, 1)
            wb.save(file)
            print("Employee deleted successfully")
            return

    print("Employee not found")

# ---------------- MAIN MENU ----------------
while True:
    print("\n===== Employee Data Automation =====")
    print("1. Add Employee")
    print("2. View Employees")
    print("3. Search Employee")
    print("4. Update Employee")
    print("5. Delete Employee")
    print("6. Exit")

    choice = input("Enter your choice: ")

    if choice == '1':
        add_employee()
    elif choice == '2':
        view_employee()
    elif choice == '3':
        search_employee()
    elif choice == '4':
        update_employee()
    elif choice == '5':
        delete_employee()
    elif choice == '6':
        print("Exiting program")
        break
    else:
        print("Invalid choice")
